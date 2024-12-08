import openpyxl

# Função para calcular o valor presente da anuidade ordinária
def anuidade_ordinaria(P, i, n):
    return P * ((1 - (1 + i)**(-n)) / i)

# Função para calcular o valor presente da anuidade vencida
def anuidade_vencida(P, i, n):
    return anuidade_ordinaria(P, i, n) * (1 + i)

# Criando uma nova planilha Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Anuidades"

# Títulos das colunas
ws['A1'] = "Pagamento (P)"
ws['B1'] = "Taxa de Juros (i)"
ws['C1'] = "Número de Períodos (n)"
ws['D1'] = "Valor Presente Anuidade Ordinária (VP)"
ws['E1'] = "Valor Presente Anuidade Vencida (VP)"

# Dados de entrada (Exemplo)
pagamento = 1000  # Pagamento periódico
taxa_juros = 0.05  # Taxa de juros de 5% por período
numero_periodos = 12  # Número de períodos (exemplo: 12 meses)

# Calculando as anuidades
vp_ordinaria = anuidade_ordinaria(pagamento, taxa_juros, numero_periodos)
vp_vencida = anuidade_vencida(pagamento, taxa_juros, numero_periodos)

# Preenchendo a planilha com os dados
ws['A2'] = pagamento
ws['B2'] = taxa_juros
ws['C2'] = numero_periodos
ws['D2'] = vp_ordinaria
ws['E2'] = vp_vencida

# Salvando o arquivo Excel
wb.save("anuidades.xlsx")

print("Planilha 'anuidades.xlsx' criada com sucesso!")
